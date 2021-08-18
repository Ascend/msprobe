# -*- coding: UTF-8 -*-
import datetime
import pandas as pd
import time
import xlwt
import xlrd
import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import openpyxl
from shutil import copyfile


today = []
target = []
# 从csv文件中读取各用例相关数据
def read_data_from_csv():
    global today
    today_tmp = datetime.date.today()
    today = str(today_tmp).replace('-', '')
    result = today + '.csv'
    # 读取所有性能数据并合并至一个csv
    for i in range(1, 5):
        file_tmp = "/root/lava_workspace/worker-10-151-245-191-" + str(i) + "/git/c3x-platform-tools/davinci-c60_aiserver-ascendDK/cloud_performance/"
        # 打印该目录下所有文件夹和文件
        files = os.listdir(file_tmp)
        # 获取今日文件名
        file_name_tmp = "cloud_performance_" + today
        for f in files:
            if file_name_tmp in f and '.csv' in f:
                file_name = file_tmp + f
        fr = open(file_name, 'rb').read()
        with open(result, 'ab') as f:
            f.write(fr)
    # 去除重复表头
    df = pd.read_csv(result, header=None)
    datalist = df.drop_duplicates()
    datalist.to_csv(result, index=False, header=False)
    data = pd.read_csv(result)
    return data

# 读取用例配置文件
def read_config_from_csv():
    data = pd.read_csv('config.csv')
    return data

# 拷贝模板文件
def copy_template():
    global target
    source = 'C75B050网络性能看护日报.xlsx'
    target = 'C75B050网络性能看护日报_' + today + '.xlsx'
    copyfile(source, target)

# 写入今日模板文件
def write_template(data, config):
    # 打开excel模板
    wb = load_workbook(target)
    # 打开sheet
    ws = wb.get_sheet_by_name('网络验收进展')
    # 异常数据标记为红色
    font = Font(u'宋体', size=10, color="FF0000")
    for i in range(len(data)):
        for j in range(len(config)):
            # 根据caseName获取数据及配置
            if data.loc[i, 'CaseName'] == config.loc[j, 'CaseName']:
                print(data.loc[i, 'CaseName'])
                print(config.loc[j, 'CaseName'])
                print(config.loc[j, 'X'])
                print(config.loc[j, 'Y'])
                print(data.loc[i, 'TrainingTime'])
                # 更新excel模板
                location1 = str(config.loc[j, 'Y']) + str(config.loc[j, 'X'])
                location2 = str(config.loc[j, 'Z']) + str(config.loc[j, 'X'])
                ws[location1] = data.loc[i, 'TrainingTime']
                ws[location2] = data.loc[i, 'ActualFPS']
                # 异常数据标红
                if data.loc[i, 'ActualFPS'] < data.loc[i, 'ExpectFPS']:
                    ws[location1].font = font
                    ws[location2].font = font
    # 保存excel文件
    wb.save(target)

if __name__ == '__main__':
    data = read_data_from_csv()
    config = read_config_from_csv()
    copy_template()
    print(data)
    print(config)
    write_template(data, config)


