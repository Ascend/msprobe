# coding: utf-8
import openpyxl as xl
import copy
from json import dumps as json_dumps


"""
@function:
    generate analysis result.
@author: 
    l00395915
@email:
    linjiashu@huawei.com
"""

unit = 100000
top_costly_num = 10
top_pending_num = 10


class Run:
    def __init__(self, name):
        self.name = name
        self.tasks = []
        self.streams = {}
        self.start_ts = 0
        self.stop_ts = 0
        self.duration = 0
        self.aic_kernel_time = 0
        self.cpu_kernel_time = 0
        self.ar_kernel_time = 0
        self.aic_exec_time = 0
        self.aic_idle_time = 0
        self.aic_idles = []
        self._json_stream = open(name + '.json', "w")
        self._statistics_stream = open(name + '.statistics', "w")
        self.ar_file = ""
        self.aic_file = ""
        self.cpu_file = ""

    def addTask(self, task):
        self.tasks.append(task)
        if task.stid in self.streams:
            st = self.streams[task.stid]
        else:
            st = Stream(task.stid)
            self.streams[task.stid] = st
        st.addTask(task)
        if self.start_ts == 0 or self.start_ts > task.start_ts:
            self.start_ts = task.start_ts
        if self.stop_ts < task.stop_ts:
            self.stop_ts = task.stop_ts
        self.duration = self.stop_ts - self.start_ts
        if task.type == 'aicore':
            self.aic_kernel_time = self.aic_kernel_time + task.duration
        elif task.type == 'aicpu':
            self.cpu_kernel_time = self.cpu_kernel_time + task.duration
        else:
            self.ar_kernel_time = self.ar_kernel_time + task.duration

    def sortTaskByTid(self, reverse=False):
        self.tasks.sort(key=lambda x: x.tid, reverse=reverse)

    def sortTaskByStid(self, reverse=False):
        self.tasks.sort(key=lambda x: x.stid, reverse=reverse)

    def sortTaskByKname(self, reverse=False):
        self.tasks.sort(key=lambda x: x.kname, reverse=reverse)

    def sortTaskByStartTs(self, reverse=False):
        self.tasks.sort(key=lambda x: x.start_ts, reverse=reverse)

    def sortTaskByStopTs(self, reverse=False):
        self.tasks.sort(key=lambda x: x.stop_ts, reverse=reverse)

    def sortTaskByDuration(self, reverse=False):
        self.tasks.sort(key=lambda x: x.duration, reverse=reverse)

    def printStatisticsToFile(self):
        self._statistics_stream.write("\n\nanalyze : %s\n" % self.name)
        self._statistics_stream.write("aic file    : %s\n" % self.aic_file)
        self._statistics_stream.write("ar file     : %s\n" % self.ar_file)
        self._statistics_stream.write("\n=========== iteration summary ========\n\n"
                                      "kernel cnt       : %d\n"
                                      "wall time        : %10f\n"
                                      "aicore busy      : %10f\n"
                                      "aicore idle      : %10f\n"
                                      "kernel time(aic) : %10f\n"
                                      "kernel time(ar)  : %10f\n" %
                                      (len(self.tasks),
                                       self.duration,
                                       self.aic_exec_time,
                                       self.aic_idle_time,
                                       self.aic_kernel_time,
                                       self.ar_kernel_time
                                       ))
        self.sortTaskByDuration(reverse=True)
        self._statistics_stream.write("========= top costly kernels ========\n")
        self._statistics_stream.write("exec time        kernel name\n")
        self._statistics_stream.write("---------        -----------\n")
        for i in range(top_costly_num):
            self._statistics_stream.write("%f \t %s\n" % (self.tasks[i].duration, self.tasks[i].kname))
        self._statistics_stream.write("\n======== top pending kernels ========\n")
        self._statistics_stream.write("pend time        kernel name\n")
        self._statistics_stream.write("---------        -----------\n")
        for i in range(top_pending_num):
            self._statistics_stream.write("%f \t %s\n" % (self.aic_idles[i][1], self.aic_idles[i][0]))
        self._statistics_stream.flush()
        self._statistics_stream.close()

    def genTimeline(self):
        self._json_stream.write(json_dumps({
            "otherData": {},
            "displayTimeUnit": "ms",
            "traceEvents": []
        })[:-2])
        self._json_stream.flush()
        for i in range(0, len(self.tasks)):
            t = self.tasks[i]
            if t.op_info:
                self._json_stream.write(json_dumps({
                    "name": t.kname,
                    "cat": t.type,
                    "ph": "X",
                    "pid": t.type,
                    "tid": t.stid,
                    "ts": t.start_ts * unit / 100,
                    "dur": t.duration * unit / 100,
                    "args": {"snapshot": {"aicore_description": t.op_info}}
                }))
            else:
                self._json_stream.write(json_dumps({
                    "name": t.kname,
                    "cat": t.type,
                    "ph": "X",
                    "pid": t.type,
                    "tid": t.stid,
                    "ts": t.start_ts * unit / 100,
                    "dur": t.duration * unit / 100
                }))
            if i < (len(self.tasks) - 1):
                self._json_stream.write(",")
            self._json_stream.flush()
        self._json_stream.write("]}")
        self._json_stream.flush()
        self._json_stream.close()

    def analyse(self):
        for t in self.tasks:
            t.start_ts = t.start_ts - self.start_ts
            t.stop_ts = t.stop_ts - self.start_ts
        self.sortTaskByStartTs()
        t_ = copy.deepcopy(self.tasks[0])
        new = True
        for t in self.tasks:
            if t.type != 'aicore':
                continue
            if new:
                self.aic_exec_time = self.aic_exec_time + t.duration
                new = False
            if t.start_ts < t_.stop_ts:
                if t.stop_ts < t_.stop_ts:
                    continue
                else:
                    self.aic_exec_time = self.aic_exec_time + t.stop_ts - t_.stop_ts
                    t_.stop_ts = t.stop_ts
                    continue
            else:
                self.aic_exec_time = self.aic_exec_time + t.duration
                self.aic_idles.append([t.kname, t.start_ts - t_.stop_ts])
                self.aic_idle_time = self.aic_idle_time + t.start_ts - t_.stop_ts
                t_ = copy.deepcopy(t)
        self.aic_idles.sort(key=lambda x: x[1], reverse=True)
        for s in self.streams.values():
            s.analyse(self.start_ts)
        self.genTimeline()


class Stream:
    def __init__(self, stid):
        self.stid = stid
        self.tasks = []
        self.start_ts = 0
        self.stop_ts = 0
        self.duration = 0
        self.aic_exec_time = 0
        self.aic_idle_time = 0

    def addTask(self, task):
        self.tasks.append(task)
        if self.start_ts == 0 or self.start_ts > task.start_ts:
            self.start_ts = task.start_ts
        if self.stop_ts < task.stop_ts:
            self.stop_ts = task.stop_ts
        self.duration = self.stop_ts - self.start_ts
        self.aic_exec_time = self.aic_exec_time + task.duration

    def sortByTid(self, reverse=False):
        self.tasks.sort(key=lambda x: x.tid, reverse=reverse)

    def sortByStid(self, reverse=False):
        self.tasks.sort(key=lambda x: x.stid, reverse=reverse)

    def sortByKname(self, reverse=False):
        self.tasks.sort(key=lambda x: x.kname, reverse=reverse)

    def sortByStartTs(self, reverse=False):
        self.tasks.sort(key=lambda x: x.start_ts, reverse=reverse)

    def sortByStopTs(self, reverse=False):
        self.tasks.sort(key=lambda x: x.stop_ts, reverse=reverse)

    def sortByDuration(self, reverse=False):
        self.tasks.sort(key=lambda x: x.duration, reverse=reverse)

    def analyse(self, start_ts):
        for t in self.tasks:
            t.start_ts = t.start_ts - start_ts
            t.stop_ts = t.start_ts - start_ts
        self.start_ts = self.start_ts - start_ts
        self.stop_ts = self.stop_ts - start_ts


class Task:
    def __init__(self, sn, tid, stid, kname, start_ts, stop_ts, type, op_info=None):
        self.sn = sn
        self.tid = tid
        self.stid = stid
        self.kname = kname
        self.start_ts = start_ts
        self.stop_ts = stop_ts
        self.duration = stop_ts - start_ts
        self.type = type
        self.op_info = op_info

    def getDuration(self):
        return self.stop_ts - self.start_ts


# @pysnooper.snoop()
def parse_aic_profiling(run, file_name, enable_op_info=False, lines=100000):
    run.aic_file = file_name
    wb = xl.load_workbook(file_name)
    ws = wb['aicore op dashboard']
    parsed_line_num = 0
    for x in range(2, lines + 1):
        if ws.cell(row=x, column=1).value is None:
            break
        sn = ws.cell(row=x, column=1).value
        tid = ws.cell(row=x, column=2).value
        stid = ws.cell(row=x, column=3).value
        kname = ws.cell(row=x, column=4).value
        start_ts = float(ws.cell(row=x, column=6).value) / unit
        stop_ts = float(ws.cell(row=x, column=7).value) / unit
        if enable_op_info:
            block_dim = ws.cell(row=x, column=9).value
            input_shape = ws.cell(row=x, column=10).value
            input_dtype = ws.cell(row=x, column=11).value
            input_format = ws.cell(row=x, column=12).value
            output_shape = ws.cell(row=x, column=13).value
            output_dtype = ws.cell(row=x, column=14).value
            output_format = ws.cell(row=x, column=15).value
            op_info = "block_dim: %s\n input_shape: {%s}\n input_dtype: %s\n input_format: %s\n" \
                      " output_shape: {%s}\n output_dtype: %s\n output_format: %s\n"\
                      % (block_dim, input_shape, input_dtype, input_format,
                         output_shape, output_dtype, output_format)
            task = Task(sn, tid, stid, kname, start_ts, stop_ts, 'aicore', op_info)
        else:
            task = Task(sn, tid, stid, kname, start_ts, stop_ts, 'aicore')
        run.addTask(task)
        parsed_line_num = parsed_line_num + 1
    return run


# @pysnooper.snoop()
def parse_ar_profiling(run, file_name, device_id, line=3):
    run.ar_file = file_name
    wb = xl.load_workbook(file_name)
    ws = wb['device %s' % device_id]
    sn = 'n/a'
    tid = 'reduce'
    stid = 'reduce'
    r = line
    c = 3
    ar_num = 0
    for i in range(1, 32):
        cell = ws.cell(row=1, column=i)
        if cell.value is not None and cell.value.startswith('Reduceadd'):
            ar_num += 1
    parsed_line_num = 0
    for i in range(1, ar_num + 1):
        if ws.cell(row=r, column=c).value is None or ws.cell(row=r, column=c + 1).value is None:
            break
        kname = 'AR' + str(i)
        start_ts = float(ws.cell(row=r, column=c).value) / unit
        c = c + 1
        stop_ts = float(ws.cell(row=r, column=c).value) / unit
        c = c + 1
        task = Task(sn, tid, stid, kname, start_ts, stop_ts, 'all_reduce')
        run.addTask(task)
        parsed_line_num += 1
    return run


# @pysnooper.snoop()
def parse_cpu_profiling(run, file_name, lines=100000):
    run.aic_file = file_name
    wb = xl.load_workbook(file_name)
    ws = wb['aicpu op dashboard']
    parsed_line_num = 0
    tid = 'aicpu'
    stid = 'aicpu'
    for x in range(2, lines + 1):
        if ws.cell(row=x, column=1).value is None:
            break
        sn = ws.cell(row=x, column=1).value
        kname = ws.cell(row=x, column=2).value
        start_ts = float(ws.cell(row=x, column=5).value) * 100 / unit
        stop_ts = float(ws.cell(row=x, column=9).value) * 100 / unit
        task = Task(sn, tid, stid, kname, start_ts, stop_ts, 'aicpu')
        run.addTask(task)
        parsed_line_num = parsed_line_num + 1
    return run


def gen_analyze(cur_time, device_id, aic_file_name, ar_file_name, aicpu_file, enable_aicpu, enable_op_info):
    run = Run('%s_training_analyze_tag.%s' % (cur_time, device_id))
    run = parse_aic_profiling(run, aic_file_name, enable_op_info)
    run = parse_ar_profiling(run, ar_file_name, device_id)
    if enable_aicpu:
        run = parse_cpu_profiling(run, aicpu_file)
    run.analyse()
    run.printStatisticsToFile()
